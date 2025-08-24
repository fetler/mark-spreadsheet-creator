import tkinter as tk
from tkinter import ttk, messagebox
from pathlib import Path
from openpyxl import load_workbook

def export_oneuni_rows_to_xlsx(
    rows,
    target_filename="/Users/mattpsychology/Documents/Marks Inputter/SEPS Master BB to OneUni mark import template - Final with IDs and marks.xlsx",
    sheet_name="Tab 3 OneUni Export",
    start_row=3,
    header_row=2,
):
    """
    Write the list[dict] produced by dnd_gui (SSPASSESS only) into the given workbook.
    - We read the headers on `header_row` of `sheet_name` and match by name (case-insensitive).
    - `start_row` is the first row for data (row 3 as requested).
    - Existing data in the target area is cleared just before writing (limited to used columns).
    """

    if not rows:
        raise ValueError("No rows to export.")

    base_dir = Path(__file__).resolve().parent
    xlsx_path = base_dir / target_filename
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Target workbook not found: {xlsx_path}")

    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"Worksheet '{sheet_name}' not found. Available: {wb.sheetnames!r}"
        )
    ws = wb[sheet_name]

    # ---- 1) Read headers on header_row and build column map
    # Example: A2, B2, C2... contain headers for Tab 3
    header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False))[0]
    headers = [ (idx+1, (cell.value or "").strip()) for idx, cell in enumerate(header_cells) ]
    used_cols = [col for col, name in headers if name]  # only columns that have header text

    # ---- 2) Map known Tab 3 header names to the dict keys coming from dnd_gui
    # NOTE: Adjust/extend this mapping to match your actual Tab 3 headers.
    # Keys = header text in workbook, Values = key in the dnd_gui row dict
    # The dict keys below match your CSV_FIELD_MAP from dnd_gui.py
    header_to_row_key = {
        "LineType": "LineType",
        "StudentStudyItemAssessmentCurriculumItemCode": "StudentStudyItemAssessmentCurriculumItemCode",
        "StudentStudyItemAssessmentCurriculumItemVersionNumber": "StudentStudyItemAssessmentCurriculumItemVersionNumber",
        "StudentStudyItemAssessmentCurriculumItemFullTitle": "StudentStudyItemAssessmentCurriculumItemFullTitle",
        "StudentStudyItemAssessmentDeliveryYear": "StudentStudyItemAssessmentDeliveryYear",
        "StudentStudyItemAssessmentDeliveryStudyPeriodCode": "StudentStudyItemAssessmentDeliveryStudyPeriodCode",
        "StudentStudyItemAssessmentDeliveryStudyPeriodDescription": "StudentStudyItemAssessmentDeliveryStudyPeriodDescription",
        "StudentStudyItemAssessmentDeliveryLocationCode": "StudentStudyItemAssessmentDeliveryLocationCode",
        "StudentStudyItemAssessmentDeliveryLocationDescription": "StudentStudyItemAssessmentDeliveryLocationDescription",
        "StudentStudyItemAssessmentDeliveryNumber": "StudentStudyItemAssessmentDeliveryNumber",
        "StudentStudyItemAssessmentStudentID": "StudentStudyItemAssessmentStudentID",
        "StudentStudyItemAssessmentStudentStudyItemAttemptNumber": "StudentStudyItemAssessmentStudentStudyItemAttemptNumber",
        "StudentStudyItemAssessmentID": "StudentStudyItemAssessmentID",
        "StudentStudyItemAssessmentTypeDescription": "StudentStudyItemAssessmentTypeDescription",
        "StudentStudyItemAssessmentDescription": "StudentStudyItemAssessmentDescription",
        "StudentStudyItemAssessmentBarcode": "StudentStudyItemAssessmentBarcode",
    }

    # case-insensitive lookup of header names
    header_lc_to_target = {h.lower(): k for h, k in header_to_row_key.items()}

    # Build final column->row_key mapping for the columns that exist on the sheet
    col_to_row_key = {}
    for col_idx, name in headers:
        name_norm = (name or "").strip().lower()
        if name_norm in header_lc_to_target:
            col_to_row_key[col_idx] = header_lc_to_target[name_norm]

    if not col_to_row_key:
        raise RuntimeError(
            "Could not match any Tab 3 headers to dnd_gui row fields.\n"
            "Please update 'header_to_row_key' to match your worksheet headers."
        )

    # ---- 3) Clear existing data region from start_row downwards for the used columns
    max_rows_to_clear = max(ws.max_row - start_row + 1, len(rows))
    for r in range(start_row, start_row + max_rows_to_clear):
        for c in used_cols:
            ws.cell(row=r, column=c, value=None)

    # ---- 4) Write data
    r = start_row
    for row_dict in rows:
        for col_idx, row_key in col_to_row_key.items():
            value = (row_dict.get(row_key, "") or "").strip()
            ws.cell(row=r, column=col_idx, value=value)
        r += 1

    wb.save(xlsx_path)
    return xlsx_path
