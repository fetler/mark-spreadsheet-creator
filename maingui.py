#!/usr/bin/env python3
from helpers.studentidsandmarks import *
from helpers.oneunicsvimport import *
from helpers.idsandmarksgui import *
from helpers.dnd_gui import *
from helpers.export_oneuni import *
import tkinter as tk
from tkinter import ttk, messagebox
from pathlib import Path
from openpyxl import load_workbook


class MinimalApp(tk.Tk):
    def __init__(self):
        super().__init__()

        # ----- Basic window setup -----
        self.title("Bulk Upload Utility")
        # High-DPI friendly scaling; tweak if needed (1.25 ~ 120 DPI)
        try:
            self.tk.call("tk", "scaling", 1.25)
        except tk.TclError:
            pass

        # Choose "dark" or "light"
        self.theme_mode = "dark"

        # ----- Theming -----
        self._apply_base_theme()
        self._create_styles()

        # ----- Layout: root frame -----
        self.root_frame = ttk.Frame(self, style="App.TFrame")
        self.root_frame.grid(sticky="nsew")
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        # Center content with padding
        self.root_frame.columnconfigure(0, weight=1)
        for r in range(3):
            self.root_frame.rowconfigure(r, weight=1)

        # Title area
        title = ttk.Label(
            self.root_frame,
            text="Bulk Upload Utility",
            style="Title.TLabel",
            anchor="center",
        )
        subtitle = ttk.Label(
            self.root_frame,
            text="Create bulk upload spreadsheets for OneUni marks input",
            style="Subtitle.TLabel",
            anchor="center",
        )
        title.grid(row=0, column=0, sticky="n", pady=(40, 4))
        subtitle.grid(row=1, column=0, sticky="n", pady=(0, 20))

        # Card container
        card = ttk.Frame(self.root_frame, style="Card.TFrame", padding=24)
        card.grid(row=2, column=0, padx=24, pady=(0, 48), sticky="n")
        card.columnconfigure(0, weight=1)

        # ---- Buttons ----
        self._make_menu_button(
            card,
            "Input Student IDs and Marks",
            command=self.on_output_ids_to_console,
            row=0
        )

        self._make_menu_button(
            card,
            "Load OneUni CSV File",
            self.on_load_one_uni_csv,
            row=1
        )

        # Print IDs to console button — NOTE: now zero-arg callable
        self._make_menu_button(
            card, "Print IDs to Console",
            self.on_print_ids_to_console,   # <-- fixed
            row=2,
        )

        self._make_menu_button(
            card,
            "Export IDs & Marks to XLSX",
            self.on_export_to_xlsx,
            row=3
        )

        self._make_menu_button(
            card,
            "Export OneUni CSV Rows to XLSX",
            self.on_export_oneuni_to_xlsx,
            row=4
        )

        # Settings button
        self._make_menu_button(
            card,
            "Settings",
            self.on_settings,
            row=5
        )

        # Center the window after layout is computed
        self.after(50, self._center_window)

        # storage for last received pairs
        self.pairs = None

    # ------------------- UI helpers -------------------
    def _apply_base_theme(self):
        """Set base colors depending on theme_mode."""
        if self.theme_mode == "dark":
            self.colors = {
                "bg": "#0b1220",  # app background
                "fg": "#e5e7eb",  # primary text
                "muted": "#9aa3b2",  # secondary text
                "card": "#111827",  # card background
                "btn": "#1f2937",  # button background
                "btn_hover": "#374151",  # button hover
                "accent": "#2563eb",  # primary action
                "accent_hover": "#1d4ed8",
                "border": "#1f2937",
            }
        else:
            self.colors = {
                "bg": "#f8fafc",
                "fg": "#0f172a",
                "muted": "#475569",
                "card": "#ffffff",
                "btn": "#f1f5f9",
                "btn_hover": "#e2e8f0",
                "accent": "#2563eb",
                "accent_hover": "#1d4ed8",
                "border": "#e5e7eb",
            }
        self.configure(bg=self.colors["bg"])

    def _create_styles(self):
        style = ttk.Style(self)
        # Use a theme that allows color customization
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        c = self.colors

        # Global frame background
        style.configure("App.TFrame", background=c["bg"])
        style.configure("Card.TFrame", background=c["card"], relief="flat")

        # Labels
        style.configure(
            "Title.TLabel",
            background=c["bg"],
            foreground=c["fg"],
            font=("SF Pro Display", 26, "bold"),
        )
        style.configure(
            "Subtitle.TLabel",
            background=c["bg"],
            foreground=c["muted"],
            font=("SF Pro Display", 14),
        )

        # Buttons (menu)
        style.configure(
            "Menu.TButton",
            background=c["btn"],
            foreground=c["fg"],
            font=("SF Pro Display", 14),
            padding=(16, 12),
            borderwidth=0,
            focuscolor=c["accent"],
            anchor="center",
        )
        style.map(
            "Menu.TButton",
            background=[("active", c["btn_hover"]), ("pressed", c["btn_hover"])],
            foreground=[("disabled", c["muted"])],
        )

        # Accent (primary) button
        style.configure(
            "Accent.TButton",
            background=c["accent"],
            foreground="#ffffff",
            font=("SF Pro Display", 14, "bold"),
            padding=(16, 12),
            borderwidth=0,
            anchor="center",
        )
        style.map(
            "Accent.TButton",
            background=[
                ("active", c["accent_hover"]),
                ("pressed", c["accent_hover"]),
            ]
        )

    def _make_menu_button(self, parent, text, command, row, accent=False):
        style = "Accent.TButton" if accent else "Menu.TButton"
        btn = ttk.Button(parent, text=text, command=command, style=style, cursor="hand2")
        btn.grid(row=row, column=0, sticky="ew", pady=(0 if row == 0 else 10, 0))
        parent.columnconfigure(0, weight=1)
        return btn

    def _center_window(self, width=620, height=580):
        # Compute a nice centered geometry
        self.update_idletasks()
        screen_w = self.winfo_screenwidth()
        screen_h = self.winfo_screenheight()
        x = int((screen_w / 2) - (width / 2))
        y = int((screen_h / 2) - (height / 2.4))
        self.geometry(f"{width}x{height}+{x}+{y}")
        self.minsize(420, 360)

    # --------------- Button callbacks ---------------
    def on_output_ids_to_console(self):
        # Open the grid window; it will call self.handle_pairs(pairs)
        app = GridApp(callback=self.handle_pairs)
        app.mainloop()

    def on_load_one_uni_csv(self):
        # Open the DnDApp with a callback to receive rows
        # Option A: open a separate window and let the user click "Send to Main"
        from helpers.dnd_gui import DnDApp
        win = DnDApp(callback=self.handle_oneuni_rows, auto_send=False)
        # If you prefer modal-like behaviour, you can just leave it to run.

    def handle_oneuni_rows(self, rows):
        """
        Receives a list[dict] of SSPASSESS rows from DnDApp.
        Store them; you can also surface a summary or enable downstream actions.
        """
        # Keep latest
        self.oneuni_rows = rows

        # Example: print a summary to console and show a messagebox
        print(f"Received {len(rows)} SSPASSESS row(s) from DnD GUI.")
        # Maybe preview first few rows:
        for i, r in enumerate(rows[:5]):
            print(f"[{i+1}] StudentID={r.get('StudentStudyItemAssessmentStudentID','')}, "
                  f"Type={r.get('StudentStudyItemAssessmentTypeDescription','')}, "
                  f"Desc={r.get('StudentStudyItemAssessmentDescription','')}")

        messagebox.showinfo(
            "CSV Loaded",
            f"Received {len(rows)} SSPASSESS row(s) from CSV."
            )
        

    def on_export_to_xlsx(self):
        """
        Exports the most recently received (id, mark) pairs into the Excel template
        and saves a 'with IDs and marks' copy alongside the template.
        """
        if not self.pairs:
            from tkinter import messagebox
            messagebox.showinfo(
                "Export to XLSX",
                "No IDs/marks to export yet.\n\n"
                "Open 'Input Student IDs and Marks, enter/paste your data, "
                "then click 'Save' in that window to send the pairs back here."
            )
            return

        try:
            out_path = export_ids_marks_to_xlsx(self.pairs)
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Export failed", str(e))
            return

        from tkinter import messagebox
        messagebox.showinfo("Export complete", f"Saved:\n{out_path}")

    def on_export_oneuni_to_xlsx(self):
        """
        Export the most recently received OneUni rows (from DnDApp) into
        'Tab 3 OneUni Export' of the 'Final - with IDs and marks' workbook,
        starting at row 3.
        """
        try:
            rows = getattr(self, "oneuni_rows", None)
            if not rows:
                from tkinter import messagebox
                messagebox.showinfo(
                    "Export OneUni Rows",
                    "No OneUni rows captured yet.\n\n"
                    "Click 'Load OneUni CSV File', drop a CSV, then click 'Send to Main'."
                )
                return

            out_path = export_oneuni_rows_to_xlsx(
                rows,
                target_filename="/Users/mattpsychology/Documents/Marks Inputter/SEPS Master BB to OneUni mark import template - Final with IDs and marks.xlsx",
                sheet_name="Tab 3 OneUni Export",
                start_row=3,
                header_row=2,
            )

            from tkinter import messagebox
            messagebox.showinfo(
                "Export Complete",
                f"Exported {len(rows)} row(s) to:\n{out_path}\n\n"
                "Worksheet: 'Tab 3 OneUni Export' (from row 3)."
            )
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Export failed", str(e))

    def on_settings(self):
        # Minimal Settings dialog (placeholder)
        win = tk.Toplevel(self)
        win.title("Settings")
        win.transient(self)
        win.configure(bg=self.colors["bg"])
        win.resizable(False, False)

        frm = ttk.Frame(win, style="App.TFrame", padding=16)
        frm.grid(sticky="nsew")
        frm.columnconfigure(0, weight=1)

        lbl = ttk.Label(
            frm, text="Settings", style="Title.TLabel", anchor="w"
        )
        lbl.grid(row=0, column=0, sticky="w", pady=(0, 10))

        desc = ttk.Label(
            frm,
            text="(Placeholder) You can add preferences here.\n"
                 "Tip: Toggle theme for a different look.",
            style="Subtitle.TLabel",
            anchor="w",
            justify="left",
        )
        desc.grid(row=1, column=0, sticky="w", pady=(0, 12))

        # Theme toggle
        toggle_text = tk.StringVar(value="Switch to Light Theme" if self.theme_mode == "dark" else "Switch to Dark Theme")

        def toggle_theme():
            self.theme_mode = "light" if self.theme_mode == "dark" else "dark"
            toggle_text.set("Switch to Light Theme" if self.theme_mode == "dark" else "Switch to Dark Theme")
            self._apply_base_theme()
            self._create_styles()

        toggle_btn = ttk.Button(frm, textvariable=toggle_text, style="Menu.TButton", command=toggle_theme, cursor="hand2")
        toggle_btn.grid(row=2, column=0, sticky="ew")

        # Center settings window relative to main
        self._center_child(win, w=420, h=220)

    def _center_child(self, win, w=420, h=220):
        self.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width() - w) // 2
        y = self.winfo_rooty() + (self.winfo_height() - h) // 2
        win.geometry(f"{w}x{h}+{x}+{y}")

    # --------- Data handoff & printing ----------
    def handle_pairs(self, pairs):
        """
        Receives (id, mark) pairs from the grid window.
        Store them for later printing. We don't print here to keep the
        'Print IDs to Console' button meaningful.
        """
        self.pairs = pairs

    def on_print_ids_to_console(self):
        """
        Zero-argument Tkinter callback for the 'Print IDs to Console' button.
        Prints the most recently received pairs to stdout.
        """
        if not self.pairs:
            messagebox.showinfo(
                "Print IDs to Console",
                "No IDs to print yet.\n\nOpen 'Input Student IDs and Marks' and "
                "click 'Output IDs and Marks' first."
            )
            return

        for sid, mark in self.pairs:
            print(f"{sid} {mark}")

def export_ids_marks_to_xlsx(
    pairs,
    template_filename="SEPS Master BB to OneUni mark import template - Final.xlsx",
    output_filename="SEPS Master BB to OneUni mark import template - Final with IDs and marks.xlsx",
    sheet_name="Tab 1 BB Export",
    start_row=1,
    ):
    """
    Write (student_id, mark) pairs to the given Excel template and save as a new file.
    - IDs go to column A (written as text to preserve leading zeros)
    - Marks go to column B (numeric if possible, otherwise text)
    - Data starts at `start_row`
    """

    base_dir = Path(__file__).resolve().parent
    template_path = base_dir / template_filename
    output_path = base_dir / output_filename

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"Worksheet '{sheet_name}' not found in template. "
            f"Available: {wb.sheetnames!r}"
            )
    ws = wb[sheet_name]

    # Optional: clear existing data region (A/B) where we'll write
    max_rows_to_clear = max(len(pairs), ws.max_row - start_row + 1)
    for r in range(start_row, start_row + max_rows_to_clear):
        ws.cell(row=r, column=1, value=None)
        ws.cell(row=r, column=2, value=None)

    r = start_row
    for sid, mark in pairs:
        # Column A: Student ID as TEXT (to preserve leading zeros)
        c_id = ws.cell(row=r, column=1, value=str(sid).strip())
        c_id.number_format = "@" # force text

        # Column B: MArk (try numeric; fallback to text)
        m_str = str(mark).strip()
        if m_str == "":
            ws.cell(row=r, column=2, value=None)
        else:
            try:
                num = float(m_str)
                # store as integer if it is an integer value like "70" or "70.0"
                if num.is_integer():
                    num = int(num)
                ws.cell(row=r, column=2, value=num)
            except ValueError:
                c_m = ws.cell(row=r, column=2, value=m_str)
                c_m.number_format = "@"

        r += 1

    wb.save(output_path)
    return output_path

if __name__ == "__main__":
    app = MinimalApp()
    app.mainloop()
